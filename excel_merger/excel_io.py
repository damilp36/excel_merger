"""Memory-conscious Excel readers used by the Streamlit app.

Modern Excel files are opened with openpyxl's read-only mode and consumed one
row at a time. The selected worksheet is materialized as a DataFrame only after
streaming because the lookup operation itself needs random access to the keys.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Callable, Iterator, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
ProgressCallback = Callable[[int, int | None], None]


class WorkbookReadError(ValueError):
    """Raised when an uploaded workbook cannot be read safely."""


@dataclass(frozen=True)
class HeaderChange:
    position: int
    original: str
    replacement: str
    reason: str


@dataclass
class LoadedSheet:
    file_name: str
    sheet_name: str
    header_row: int
    dataframe: pd.DataFrame
    header_changes: list[HeaderChange]
    blank_rows_removed: int = 0
    blank_columns_removed: int = 0


def _extension(file_name: str) -> str:
    extension = Path(file_name).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise WorkbookReadError(
            f"{file_name!r} is not a supported Excel file. Use .xlsx, .xlsm, or .xls."
        )
    return extension


def _friendly_error(file_name: str, exc: Exception) -> WorkbookReadError:
    if isinstance(exc, ImportError):
        return WorkbookReadError(
            f"Reading legacy .xls files requires the optional 'xlrd' package. "
            f"Convert {file_name!r} to .xlsx or install the project requirements."
        )
    return WorkbookReadError(
        f"Could not read {file_name!r}. Confirm that it is a valid, unencrypted Excel workbook."
    )


def list_sheet_names(content: bytes, file_name: str) -> list[str]:
    """Return workbook sheet names without loading worksheet data."""

    extension = _extension(file_name)
    try:
        if extension == ".xls":
            with pd.ExcelFile(BytesIO(content), engine="xlrd") as workbook:
                return list(workbook.sheet_names)

        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    except (OSError, ValueError, KeyError, InvalidFileException, ImportError) as exc:
        raise _friendly_error(file_name, exc) from exc


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        missing = pd.isna(value)
    except (TypeError, ValueError):
        return False
    return bool(missing) if not hasattr(missing, "__len__") else False


def _trim_trailing_empty(values: Sequence[Any]) -> list[Any]:
    result = list(values)
    while result and _is_missing(result[-1]):
        result.pop()
    return result


def _normalise_headers(values: Sequence[Any]) -> tuple[list[str], list[HeaderChange]]:
    headers: list[str] = []
    changes: list[HeaderChange] = []
    counts: dict[str, int] = {}

    for index, raw_value in enumerate(values, start=1):
        raw_text = "" if _is_missing(raw_value) else str(raw_value).strip()
        # pandas synthesizes these labels when reading blank headers in .xls files.
        if re.fullmatch(r"Unnamed:\s*\d+", raw_text):
            raw_text = ""
        base = raw_text or f"Unnamed_{index}"
        if not raw_text:
            changes.append(
                HeaderChange(index, "(blank)", base, "Blank header was given a name")
            )

        counts[base] = counts.get(base, 0) + 1
        header = base if counts[base] == 1 else f"{base}__{counts[base]}"
        if header != base:
            changes.append(
                HeaderChange(index, base, header, "Duplicate header was made unique")
            )
        headers.append(header)

    return headers, changes


def _stream_openpyxl_rows(
    content: bytes,
    sheet_name: str,
) -> tuple[Iterator[tuple[Any, ...]], int | None, Any]:
    workbook = load_workbook(
        BytesIO(content), read_only=True, data_only=True, keep_links=False
    )
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise WorkbookReadError(f"Worksheet {sheet_name!r} no longer exists in this workbook.")
    worksheet = workbook[sheet_name]
    return worksheet.iter_rows(values_only=True), worksheet.max_row, workbook


def stream_sheet(
    content: bytes,
    file_name: str,
    sheet_name: str,
    header_row: int = 1,
    progress: ProgressCallback | None = None,
) -> LoadedSheet:
    """Stream a worksheet into a normalized DataFrame.

    ``header_row`` is one-based. Fully empty trailing rows and columns are
    removed, while intentional blank rows inside the data are preserved.
    """

    if header_row < 1:
        raise WorkbookReadError("Header row must be 1 or greater.")

    extension = _extension(file_name)
    if extension == ".xls":
        try:
            dataframe = pd.read_excel(
                BytesIO(content),
                sheet_name=sheet_name,
                header=header_row - 1,
                engine="xlrd",
                dtype=object,
            )
        except (OSError, ValueError, KeyError, ImportError) as exc:
            raise _friendly_error(file_name, exc) from exc
        original_headers = list(dataframe.columns)
        headers, changes = _normalise_headers(original_headers)
        dataframe.columns = headers
        empty_columns = [column for column in dataframe if dataframe[column].isna().all()]
        removable_columns = [
            column for column in empty_columns if column.startswith("Unnamed_")
        ]
        dataframe = dataframe.drop(columns=removable_columns)
        blank_rows = int(dataframe.isna().all(axis=1).sum())
        dataframe = dataframe.dropna(axis=0, how="all").reset_index(drop=True)
        return LoadedSheet(
            file_name,
            sheet_name,
            header_row,
            dataframe,
            changes,
            blank_rows_removed=blank_rows,
            blank_columns_removed=len(removable_columns),
        )

    workbook = None
    try:
        rows, total_rows, workbook = _stream_openpyxl_rows(content, sheet_name)
        header_values: list[Any] | None = None
        streamed_rows: list[list[Any]] = []

        for row_number, row in enumerate(rows, start=1):
            if progress and (row_number == 1 or row_number % 1_000 == 0):
                progress(row_number, total_rows)
            if row_number < header_row:
                continue
            values = _trim_trailing_empty(row)
            if row_number == header_row:
                header_values = values
                continue
            streamed_rows.append(values)

        if progress:
            progress(total_rows or len(streamed_rows) + header_row, total_rows)
        if header_values is None:
            raise WorkbookReadError(
                f"Header row {header_row} is outside worksheet {sheet_name!r}."
            )
        if not header_values:
            raise WorkbookReadError(
                f"Header row {header_row} in worksheet {sheet_name!r} is empty."
            )

        column_count = max(
            len(header_values), max((len(row) for row in streamed_rows), default=0)
        )
        padded_header = header_values + [None] * (column_count - len(header_values))
        headers, changes = _normalise_headers(padded_header)
        padded_rows = [row + [None] * (column_count - len(row)) for row in streamed_rows]

        trailing_blank_rows = 0
        while padded_rows and all(_is_missing(value) for value in padded_rows[-1]):
            padded_rows.pop()
            trailing_blank_rows += 1

        dataframe = pd.DataFrame(padded_rows, columns=headers, dtype=object)
        empty_columns = [column for column in dataframe if dataframe[column].isna().all()]
        # Keep named, intentionally empty columns; remove only generated blank columns.
        removable_columns = [
            column for column in empty_columns if column.startswith("Unnamed_")
        ]
        dataframe = dataframe.drop(columns=removable_columns)

        return LoadedSheet(
            file_name,
            sheet_name,
            header_row,
            dataframe,
            changes,
            blank_rows_removed=trailing_blank_rows,
            blank_columns_removed=len(removable_columns),
        )
    except WorkbookReadError:
        raise
    except (OSError, ValueError, KeyError, InvalidFileException) as exc:
        raise _friendly_error(file_name, exc) from exc
    finally:
        if workbook is not None:
            workbook.close()
