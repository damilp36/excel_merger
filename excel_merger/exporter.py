"""Create a polished, auditable Excel download in memory."""

from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
import math
from pathlib import Path
import re
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .merge import MergeResult
from .quality import QualityReport, is_missing


EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384

NAVY = "17324D"
BLUE = "2563EB"
PALE_RED = "FEE2E2"
PALE_AMBER = "FEF3C7"
PALE_GREEN = "DCFCE7"
PALE_GRAY = "F8FAFC"
TEXT = "172033"
WHITE = "FFFFFF"
BORDER_COLOR = "DCE4EC"


class ExportLimitError(ValueError):
    """Raised when merged data exceeds the Excel worksheet grid."""


def clean_output_filename(name: str) -> str:
    raw = Path(name.strip()).name if name.strip() else "merged_lookup.xlsx"
    stem = Path(raw).stem
    stem = re.sub(r"[\\/:*?\"<>|]+", "_", stem).strip(" ._")
    stem = stem[:140] or "merged_lookup"
    return f"{stem}.xlsx"


def _safe_value(value: Any) -> Any:
    if is_missing(value):
        return None
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except (ValueError, AttributeError):
            pass
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime) and value.tzinfo is not None:
        value = value.replace(tzinfo=None)
    if isinstance(value, (datetime, date, time, bool, int, float, str)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)[:32_767]
        return value
    if isinstance(value, Decimal):
        return float(value)
    return ILLEGAL_CHARACTERS_RE.sub("", str(value))[:32_767]


def _cell(
    worksheet: Any,
    value: Any,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> WriteOnlyCell:
    safe_value = _safe_value(value)
    cell = WriteOnlyCell(worksheet, value=safe_value)
    if isinstance(safe_value, str) and safe_value.startswith("="):
        cell.data_type = "s"
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def _fit_widths(dataframe: pd.DataFrame, sample_rows: int = 250) -> list[float]:
    widths: list[float] = []
    sample = dataframe.head(sample_rows)
    for column in dataframe.columns:
        lengths = [len(str(column))]
        lengths.extend(
            min(len(str(value)), 60)
            for value in sample[column]
            if not is_missing(value)
        )
        width = max(lengths, default=10) + 2
        widths.append(float(min(max(width, 10), 40)))
    return widths


def _write_merged_sheet(workbook: Workbook, dataframe: pd.DataFrame) -> None:
    if len(dataframe) + 1 > EXCEL_MAX_ROWS:
        raise ExportLimitError(
            f"The merged result has {len(dataframe):,} rows, exceeding Excel's "
            f"{EXCEL_MAX_ROWS - 1:,}-data-row worksheet limit."
        )
    if len(dataframe.columns) > EXCEL_MAX_COLUMNS:
        raise ExportLimitError(
            f"The merged result has {len(dataframe.columns):,} columns, exceeding "
            f"Excel's {EXCEL_MAX_COLUMNS:,}-column limit."
        )

    worksheet = workbook.create_sheet("Merged Data")
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(max(len(dataframe.columns), 1))}{len(dataframe) + 1}"
    )
    worksheet.row_dimensions[1].height = 28

    header_fill = PatternFill("solid", fgColor=NAVY)
    missing_fill = PatternFill("solid", fgColor=PALE_RED)
    warning_fill = PatternFill("solid", fgColor=PALE_AMBER)
    success_fill = PatternFill("solid", fgColor=PALE_GREEN)
    alternate_fill = PatternFill("solid", fgColor=PALE_GRAY)
    thin = Side(style="thin", color=BORDER_COLOR)
    bottom_border = Border(bottom=thin)
    header_font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    body_font = Font(name="Aptos", size=10, color=TEXT)

    worksheet.append(
        [
            _cell(
                worksheet,
                column,
                fill=header_fill,
                font=header_font,
                alignment=Alignment(vertical="center"),
            )
            for column in dataframe.columns
        ]
    )

    status_columns = {
        index
        for index, column in enumerate(dataframe.columns)
        if str(column).endswith(".Match status")
    }
    for row_number, values in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        output_row: list[WriteOnlyCell] = []
        for column_number, value in enumerate(values):
            safe_value = _safe_value(value)
            fill = alternate_fill if row_number % 2 == 0 else None
            font = body_font
            if safe_value is None:
                fill = missing_fill
            elif column_number in status_columns:
                if safe_value == "Matched":
                    fill = success_fill
                elif safe_value in {"Not matched", "Missing key"}:
                    fill = warning_fill
            output_row.append(
                _cell(
                    worksheet,
                    safe_value,
                    fill=fill,
                    font=font,
                    alignment=Alignment(vertical="top"),
                    border=bottom_border,
                )
            )
        worksheet.append(output_row)

    for index, width in enumerate(_fit_widths(dataframe), start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _section_title(worksheet: Any, title: str, width: int) -> None:
    title_fill = PatternFill("solid", fgColor=NAVY)
    title_font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    worksheet.append(
        [
            _cell(
                worksheet,
                title,
                fill=title_fill,
                font=title_font,
                alignment=Alignment(vertical="center"),
            ),
            *[
                _cell(worksheet, None, fill=title_fill)
                for _ in range(max(width - 1, 0))
            ],
        ]
    )


def _header_row(worksheet: Any, values: Iterable[str]) -> list[WriteOnlyCell]:
    fill = PatternFill("solid", fgColor=BLUE)
    font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    return [
        _cell(
            worksheet,
            value,
            fill=fill,
            font=font,
            alignment=Alignment(vertical="center", wrap_text=True),
        )
        for value in values
    ]


def _write_audit_sheet(workbook: Workbook, result: MergeResult) -> None:
    worksheet = workbook.create_sheet("Lookup Audit")
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A8"
    _section_title(worksheet, "Lookup Audit", 11)
    worksheet.row_dimensions[1].height = 32
    worksheet.append(["Base rows", result.base_rows])
    worksheet.append(["Output rows", len(result.dataframe)])
    worksheet.append(["Base lookup key", result.base_key])
    worksheet.append([])
    worksheet.append(
        [
            "Each lookup is measured against the original base rows. Duplicate expansion can increase output rows."
        ]
    )
    worksheet.append([])

    audit_frame = result.audit_frame()
    worksheet.append(_header_row(worksheet, [str(column) for column in audit_frame.columns]))
    body_font = Font(name="Aptos", size=10, color=TEXT)
    for row in audit_frame.itertuples(index=False, name=None):
        worksheet.append([_cell(worksheet, value, font=body_font) for value in row])

    widths = [24, 20, 20, 14, 15, 19, 18, 20, 18, 18, 17]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _write_quality_sheet(
    workbook: Workbook,
    reports: list[QualityReport],
) -> None:
    worksheet = workbook.create_sheet("Data Quality")
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A5"
    headers = [
        "File",
        "Worksheet",
        "Rows",
        "Columns",
        "Missing cells",
        "Mixed-type columns",
        "Duplicate rows",
        "Blank rows removed",
        "Blank columns removed",
        "Header fixes",
        "Findings",
    ]
    _section_title(worksheet, "Data Quality Summary", len(headers))
    worksheet.row_dimensions[1].height = 32
    worksheet.append(
        [
            "Legend",
            _cell(
                worksheet,
                "Missing values",
                fill=PatternFill("solid", fgColor=PALE_RED),
            ),
            _cell(
                worksheet,
                "Inconsistent values / warnings",
                fill=PatternFill("solid", fgColor=PALE_AMBER),
            ),
        ]
    )
    worksheet.append([])
    worksheet.append(_header_row(worksheet, headers))

    warning_fill = PatternFill("solid", fgColor=PALE_AMBER)
    body_font = Font(name="Aptos", size=10, color=TEXT)
    for report in reports:
        findings = "; ".join(report.warnings) if report.warnings else "No issues detected"
        values = [
            report.file_name,
            report.sheet_name,
            report.row_count,
            report.column_count,
            report.missing_cells,
            report.mixed_type_columns,
            report.duplicate_rows,
            report.blank_rows_removed,
            report.blank_columns_removed,
            len(report.header_changes),
            findings,
        ]
        row: list[WriteOnlyCell] = []
        for index, value in enumerate(values):
            has_warning = (
                (index in {4, 5, 6, 7, 8, 9} and isinstance(value, int) and value > 0)
                or (index == 10 and report.warnings)
            )
            row.append(
                _cell(
                    worksheet,
                    value,
                    fill=warning_fill if has_warning else None,
                    font=body_font,
                    alignment=Alignment(vertical="top", wrap_text=index == 10),
                )
            )
        worksheet.append(row)

    widths = [28, 22, 12, 12, 16, 19, 16, 20, 22, 14, 42]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def build_output_workbook(
    result: MergeResult,
    quality_reports: list[QualityReport],
) -> bytes:
    """Build the downloadable .xlsx workbook and return its bytes."""

    workbook = Workbook(write_only=True)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    _write_merged_sheet(workbook, result.dataframe)
    _write_audit_sheet(workbook, result)
    _write_quality_sheet(workbook, quality_reports)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
