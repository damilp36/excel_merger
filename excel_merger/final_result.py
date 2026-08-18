"""Final-result calculations and workbook generation.

The layout is based on a conventional academic mark sheet: course details,
grade distribution, pass statistics, remarks, and a detailed result table.
Uploaded reference values are never copied into the generated workbook.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import math
from typing import Any
from zipfile import BadZipFile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.table import Table, TableStyleInfo

from .quality import is_missing


# Reference-workbook palette. The pale beige and dark brown combination keeps
# the generated mark sheet visually consistent with the supplied template.
TEMPLATE_BEIGE = "EEECE1"
TEMPLATE_BROWN = "663300"
GRADE_GREEN = "548235"
GRADE_DARK_GREEN = "375623"
GRADE_LIGHT_GREEN = "E2F0D9"
PALE_GREEN = "DCFCE7"
PALE_RED = "FEE2E2"
PALE_AMBER = "FEF3C7"
WHITE = "FFFFFF"
TEXT = "000000"
MUTED = TEMPLATE_BROWN
BORDER_COLOR = "D9D1C5"


@dataclass(frozen=True)
class GradeScale:
    a_minimum: float = 70
    b_minimum: float = 60
    c_minimum: float = 50
    d_minimum: float = 45
    e_minimum: float = 40
    pass_mark: float = 40

    def validation_errors(self, maximum_total: float) -> list[str]:
        errors: list[str] = []
        thresholds = [
            self.a_minimum,
            self.b_minimum,
            self.c_minimum,
            self.d_minimum,
            self.e_minimum,
        ]
        if not all(
            thresholds[index] > thresholds[index + 1]
            for index in range(len(thresholds) - 1)
        ):
            errors.append("Grade minimums must descend in order from A through E.")
        if any(value < 0 or value > maximum_total for value in thresholds):
            errors.append(
                f"Every grade minimum must be between 0 and {maximum_total:g}."
            )
        if self.pass_mark < 0 or self.pass_mark > maximum_total:
            errors.append(f"The pass mark must be between 0 and {maximum_total:g}.")
        return errors

    def grade_for(self, total: float) -> tuple[str, int]:
        if total >= self.a_minimum:
            return "A", 5
        if total >= self.b_minimum:
            return "B", 4
        if total >= self.c_minimum:
            return "C", 3
        if total >= self.d_minimum:
            return "D", 2
        if total >= self.e_minimum:
            return "E", 1
        return "F", 0


@dataclass(frozen=True)
class FinalResultDetails:
    programme: str
    session: str
    semester: str
    course_code: str
    course_title: str
    credit_units: int
    course_status: str
    lecturers: str
    remarks: str
    identifier_heading: str
    maximum_ca_score: float
    maximum_exam_score: float
    grade_scale: GradeScale

    @property
    def maximum_total(self) -> float:
        return self.maximum_ca_score + self.maximum_exam_score

    def validation_errors(self) -> list[str]:
        errors: list[str] = []
        if self.maximum_ca_score <= 0:
            errors.append("Maximum CA score must be greater than zero.")
        if self.maximum_exam_score <= 0:
            errors.append("Maximum exam score must be greater than zero.")
        if self.maximum_total > 1_000:
            errors.append("The combined maximum score must not exceed 1,000.")
        if not self.identifier_heading.strip():
            errors.append("The identifier heading cannot be blank.")
        if self.maximum_total > 0:
            errors.extend(self.grade_scale.validation_errors(self.maximum_total))
        return errors


@dataclass
class ProcessedFinalResult:
    dataframe: pd.DataFrame
    grade_counts: dict[str, int]
    pass_count: int
    fail_count: int
    incomplete_count: int
    missing_ca_count: int
    missing_exam_count: int
    invalid_ca_count: int
    invalid_exam_count: int

    @property
    def complete_count(self) -> int:
        return self.pass_count + self.fail_count

    @property
    def total_count(self) -> int:
        return len(self.dataframe)

    @property
    def pass_rate(self) -> float:
        return self.pass_count / self.complete_count if self.complete_count else 0.0


@dataclass
class ImportedFinalResult:
    details: FinalResultDetails
    processed: ProcessedFinalResult
    source_dataframe: pd.DataFrame


class FinalResultWorkbookError(ValueError):
    """Raised when an uploaded workbook is not a generated final result."""


def _clean_identifier(value: Any) -> str | None:
    if is_missing(value):
        return None
    text = str(value).strip()
    return text or None


def _score_is_missing(value: Any) -> bool:
    return is_missing(value) or (isinstance(value, str) and not value.strip())


def _numeric_scores(
    series: pd.Series,
    maximum: float,
) -> tuple[pd.Series, pd.Series, pd.Series]:
    numeric = pd.to_numeric(series, errors="coerce")
    missing = series.map(_score_is_missing)
    source_has_value = ~missing
    invalid = source_has_value & (numeric.isna() | (numeric < 0) | (numeric > maximum))
    clean = numeric.mask(missing, 0).mask(invalid)
    return clean.astype("Float64"), invalid, missing


def _incomplete_status(
    ca_missing: bool,
    exam_missing: bool,
    ca_invalid: bool,
    exam_invalid: bool,
) -> str | None:
    if ca_missing and exam_missing:
        return "Incomplete: Missing CA and Exam"
    if ca_invalid and exam_invalid:
        return "Incomplete: Invalid CA and Exam"

    issues: list[str] = []
    if ca_missing:
        issues.append("Missing CA")
    elif ca_invalid:
        issues.append("Invalid CA")
    if exam_missing:
        issues.append("Missing Exam")
    elif exam_invalid:
        issues.append("Invalid Exam")
    return f"Incomplete: {' and '.join(issues)}" if issues else None


def prepare_final_result(
    merged_dataframe: pd.DataFrame,
    identifier_column: str,
    ca_column: str,
    exam_column: str,
    details: FinalResultDetails,
) -> ProcessedFinalResult:
    """Validate scores and calculate a preview of the final result."""

    errors = details.validation_errors()
    if errors:
        raise ValueError(" ".join(errors))
    selected = [identifier_column, ca_column, exam_column]
    missing_columns = [column for column in selected if column not in merged_dataframe.columns]
    if missing_columns:
        raise ValueError("Selected columns do not exist: " + ", ".join(missing_columns))
    if len(set(selected)) != 3:
        raise ValueError("Identifier, CA score, and exam score must use different columns.")

    ca_scores, invalid_ca, missing_ca = _numeric_scores(
        merged_dataframe[ca_column], details.maximum_ca_score
    )
    exam_scores, invalid_exam, missing_exam = _numeric_scores(
        merged_dataframe[exam_column], details.maximum_exam_score
    )
    identifiers = merged_dataframe[identifier_column].map(_clean_identifier)
    scores_are_valid = ca_scores.notna() & exam_scores.notna()
    totals = (ca_scores + exam_scores).where(scores_are_valid)

    grades: list[str | None] = []
    grade_points: list[int | None] = []
    statuses: list[str] = []
    row_states = zip(
        totals.tolist(),
        missing_ca.tolist(),
        missing_exam.tolist(),
        invalid_ca.tolist(),
        invalid_exam.tolist(),
        strict=True,
    )
    for total, ca_is_missing, exam_is_missing, ca_is_invalid, exam_is_invalid in row_states:
        issue_status = _incomplete_status(
            bool(ca_is_missing),
            bool(exam_is_missing),
            bool(ca_is_invalid),
            bool(exam_is_invalid),
        )
        if is_missing(total):
            grades.append(None)
            grade_points.append(None)
            statuses.append(issue_status or "Incomplete")
            continue
        grade, point = details.grade_scale.grade_for(float(total))
        grades.append(grade)
        grade_points.append(point)
        statuses.append(
            issue_status
            or ("Pass" if float(total) >= details.grade_scale.pass_mark else "Fail")
        )

    result = pd.DataFrame(
        {
            "S/No.": range(1, len(merged_dataframe) + 1),
            details.identifier_heading.strip(): identifiers,
            "CA Score": ca_scores,
            "Exam Score": exam_scores,
            "Grand Total": totals,
            "Grade Letter": grades,
            "Grade Point": grade_points,
            "Student Status": statuses,
        }
    )
    grade_counts = {
        grade: int(result["Grade Letter"].eq(grade).sum()) for grade in "ABCDEF"
    }
    return ProcessedFinalResult(
        dataframe=result,
        grade_counts=grade_counts,
        pass_count=statuses.count("Pass"),
        fail_count=statuses.count("Fail"),
        incomplete_count=sum(status.startswith("Incomplete") for status in statuses),
        missing_ca_count=int(missing_ca.sum()),
        missing_exam_count=int(missing_exam.sum()),
        invalid_ca_count=int(invalid_ca.sum()),
        invalid_exam_count=int(invalid_exam.sum()),
    )


def prepare_edited_final_result(
    edited_dataframe: pd.DataFrame,
    details: FinalResultDetails,
) -> ProcessedFinalResult:
    """Recalculate a final result after CA or Exam values are edited."""

    identifier_column = details.identifier_heading.strip()
    required_columns = [identifier_column, "CA Score", "Exam Score"]
    missing_columns = [
        column for column in required_columns if column not in edited_dataframe.columns
    ]
    if missing_columns:
        raise ValueError(
            "The editable result is missing required columns: "
            + ", ".join(missing_columns)
        )

    editable_source = edited_dataframe.copy()
    for score_column in ["CA Score", "Exam Score"]:
        editable_source[score_column] = editable_source[score_column].astype("object")
    if "Student Status" in editable_source.columns:
        for row_index, status_value in editable_source["Student Status"].items():
            status = str(status_value)
            for score_column, score_name, maximum_score in [
                ("CA Score", "CA", details.maximum_ca_score),
                ("Exam Score", "Exam", details.maximum_exam_score),
            ]:
                value = editable_source.at[row_index, score_column]
                missing_issue = _status_has_score_issue(
                    status, "Missing", score_name
                )
                invalid_issue = _status_has_score_issue(
                    status, "Invalid", score_name
                )
                if missing_issue and not is_missing(value):
                    try:
                        still_placeholder_zero = float(value) == 0
                    except (TypeError, ValueError):
                        still_placeholder_zero = False
                    if still_placeholder_zero:
                        editable_source.at[row_index, score_column] = None
                elif invalid_issue and is_missing(value):
                    # A numeric out-of-range marker preserves the invalid state
                    # without inserting text into a numeric score column.
                    editable_source.at[row_index, score_column] = maximum_score + 1

    return prepare_final_result(
        editable_source,
        identifier_column,
        "CA Score",
        "Exam Score",
        details,
    )


def _status_has_score_issue(status: str, issue: str, score_name: str) -> bool:
    return (
        f"{issue} {score_name}" in status
        or (score_name == "Exam" and f"{issue} CA and Exam" in status)
    )


def apply_bulk_score_adjustment(
    processed: ProcessedFinalResult,
    details: FinalResultDetails,
    score_column: str,
    amount: float,
    grade_letter: str | None = None,
    include_missing: bool = False,
) -> tuple[ProcessedFinalResult, int]:
    """Add marks to one score column for a grade group or all records."""

    if score_column not in {"CA Score", "Exam Score"}:
        raise ValueError("Bulk adjustments can only be applied to CA or Exam scores.")
    if amount < 0:
        raise ValueError("The number of marks to add cannot be negative.")
    if grade_letter is not None and grade_letter not in set("ABCDEF"):
        raise ValueError("The selected grade group must be A, B, C, D, E, or F.")

    edited = processed.dataframe.copy()
    current_scores = pd.to_numeric(edited[score_column], errors="coerce")
    eligible = current_scores.notna()
    if grade_letter is not None:
        eligible &= edited["Grade Letter"].eq(grade_letter)

    score_name = "CA" if score_column == "CA Score" else "Exam"
    if not include_missing:
        missing_target = edited["Student Status"].map(
            lambda status: _status_has_score_issue(
                str(status), "Missing", score_name
            )
        )
        eligible &= ~missing_target

    maximum = (
        details.maximum_ca_score
        if score_column == "CA Score"
        else details.maximum_exam_score
    )
    adjusted_scores = (current_scores + float(amount)).clip(upper=maximum)
    changed = eligible & adjusted_scores.ne(current_scores)
    edited.loc[changed, score_column] = adjusted_scores.loc[changed]
    return prepare_edited_final_result(edited, details), int(changed.sum())


def import_final_result_workbook(content: bytes) -> ImportedFinalResult:
    """Load a workbook previously created by the final-result exporter."""

    try:
        workbook = load_workbook(BytesIO(content), data_only=False)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise FinalResultWorkbookError(
            "The uploaded file is not a readable Excel workbook."
        ) from exc

    try:
        required_sheets = {"Final Result", "Settings"}
        if not required_sheets.issubset(workbook.sheetnames):
            raise FinalResultWorkbookError(
                "This is not a generated final-result workbook. It must contain "
                "Final Result and Settings worksheets."
            )

        result = workbook["Final Result"]
        settings = workbook["Settings"]
        expected_headers = {
            "A39": "S/No.",
            "C39": "CA Score",
            "D39": "Exam Score",
            "E39": "Grand Total",
            "F39": "Grade Letter",
            "G39": "Grade Point",
            "H39": "Student Status",
        }
        incorrect_headers = [
            cell_ref
            for cell_ref, expected in expected_headers.items()
            if result[cell_ref].value != expected
        ]
        identifier_heading = str(result["B39"].value or "").strip()
        if incorrect_headers or not identifier_heading:
            raise FinalResultWorkbookError(
                "The Final Result worksheet does not have the expected result table."
            )

        def setting_text(cell_ref: str) -> str:
            value = settings[cell_ref].value
            return "" if value is None else str(value)

        def setting_number(cell_ref: str, label: str) -> float:
            value = settings[cell_ref].value
            try:
                return float(value)
            except (TypeError, ValueError) as exc:
                raise FinalResultWorkbookError(
                    f"The Settings worksheet has an invalid {label}."
                ) from exc

        credit_units_value = setting_number("B8", "credit-unit value")
        if not credit_units_value.is_integer():
            raise FinalResultWorkbookError(
                "The Settings worksheet credit-unit value must be a whole number."
            )
        details = FinalResultDetails(
            programme=setting_text("B3"),
            session=setting_text("B4"),
            semester=setting_text("B5"),
            course_code=setting_text("B6"),
            course_title=setting_text("B7"),
            credit_units=int(credit_units_value),
            course_status=setting_text("B9"),
            lecturers=setting_text("B10"),
            remarks=setting_text("B11"),
            identifier_heading=identifier_heading,
            maximum_ca_score=setting_number("B14", "maximum CA score"),
            maximum_exam_score=setting_number("B15", "maximum Exam score"),
            grade_scale=GradeScale(
                a_minimum=setting_number("B20", "A-grade minimum"),
                b_minimum=setting_number("B21", "B-grade minimum"),
                c_minimum=setting_number("B22", "C-grade minimum"),
                d_minimum=setting_number("B23", "D-grade minimum"),
                e_minimum=setting_number("B24", "E-grade minimum"),
                pass_mark=setting_number("B16", "pass mark"),
            ),
        )
        validation_errors = details.validation_errors()
        if validation_errors:
            raise FinalResultWorkbookError(" ".join(validation_errors))

        if "FinalResultTable" in result.tables:
            _, _, _, data_end = range_boundaries(
                result.tables["FinalResultTable"].ref
            )
        else:
            data_end = result.max_row
        if data_end < 40:
            raise FinalResultWorkbookError(
                "The Final Result worksheet does not contain any result records."
            )

        identifiers: list[Any] = []
        ca_values: list[Any] = []
        exam_values: list[Any] = []
        for row_number in range(40, data_end + 1):
            identifier = result.cell(row_number, 2).value
            ca_value = result.cell(row_number, 3).value
            exam_value = result.cell(row_number, 4).value
            status = str(result.cell(row_number, 8).value or "")
            if all(is_missing(value) for value in [identifier, ca_value, exam_value]):
                continue
            if _status_has_score_issue(status, "Missing", "CA"):
                ca_value = None
            elif _status_has_score_issue(status, "Invalid", "CA"):
                ca_value = details.maximum_ca_score + 1
            if _status_has_score_issue(status, "Missing", "Exam"):
                exam_value = None
            elif _status_has_score_issue(status, "Invalid", "Exam"):
                exam_value = details.maximum_exam_score + 1
            identifiers.append(identifier)
            ca_values.append(ca_value)
            exam_values.append(exam_value)

        source = pd.DataFrame(
            {
                identifier_heading: identifiers,
                "CA Score": ca_values,
                "Exam Score": exam_values,
            }
        )
        if source.empty:
            raise FinalResultWorkbookError(
                "The Final Result worksheet does not contain any result records."
            )
        processed = prepare_final_result(
            source,
            identifier_heading,
            "CA Score",
            "Exam Score",
            details,
        )
        return ImportedFinalResult(
            details=details,
            processed=processed,
            source_dataframe=source,
        )
    except FinalResultWorkbookError:
        raise
    except (KeyError, TypeError, ValueError) as exc:
        raise FinalResultWorkbookError(
            "The uploaded workbook could not be read as a generated final result."
        ) from exc
    finally:
        workbook.close()


def _safe_excel_value(value: Any) -> Any:
    if is_missing(value):
        return None
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            value = value.item()
        except (TypeError, ValueError, AttributeError):
            pass
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def _set_literal(cell: Any, value: Any) -> None:
    safe = _safe_excel_value(value)
    cell.value = safe
    if isinstance(safe, str) and safe.startswith("="):
        cell.data_type = "s"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _apply_section_header(
    worksheet: Any,
    cell_range: str,
    title: str,
    fill_color: str = TEMPLATE_BEIGE,
    font_color: str = TEMPLATE_BROWN,
) -> None:
    worksheet.merge_cells(cell_range)
    cell = worksheet[cell_range.split(":", 1)[0]]
    cell.value = title
    cell.fill = _fill(fill_color)
    cell.font = Font(
        name="Aptos Display",
        size=12,
        bold=True,
        color=font_color,
    )
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _build_settings_sheet(
    workbook: Workbook,
    details: FinalResultDetails,
) -> Any:
    worksheet = workbook.create_sheet("Settings")
    worksheet.sheet_view.showGridLines = True
    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = "Final Result Settings"
    worksheet["A1"].fill = _fill(TEMPLATE_BEIGE)
    worksheet["A1"].font = Font(
        name="Aptos Display",
        size=16,
        bold=True,
        color=TEMPLATE_BROWN,
    )
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 30

    details_rows = [
        ("Programme", details.programme),
        ("Session", details.session),
        ("Semester", details.semester),
        ("Course Code", details.course_code),
        ("Course Title", details.course_title),
        ("Credit Units", details.credit_units),
        ("Course Status", details.course_status),
        ("Lecturer(s)", details.lecturers),
        ("Remarks", details.remarks),
    ]
    for row_number, (label, value) in enumerate(details_rows, start=3):
        worksheet.cell(row_number, 1, label)
        _set_literal(worksheet.cell(row_number, 2), value)

    worksheet.merge_cells("A13:D13")
    worksheet["A13"] = "Score Settings"
    worksheet["A13"].fill = _fill(TEMPLATE_BEIGE)
    worksheet["A13"].font = Font(name="Aptos", bold=True, color=TEMPLATE_BROWN)
    score_rows = [
        ("Maximum CA Score", details.maximum_ca_score),
        ("Maximum Exam Score", details.maximum_exam_score),
        ("Pass Mark", details.grade_scale.pass_mark),
    ]
    for row_number, (label, value) in enumerate(score_rows, start=14):
        worksheet.cell(row_number, 1, label)
        worksheet.cell(row_number, 2, value)

    worksheet.merge_cells("A18:D18")
    worksheet["A18"] = "Grading Scale"
    worksheet["A18"].fill = _fill(GRADE_LIGHT_GREEN)
    worksheet["A18"].font = Font(name="Aptos", bold=True, color=GRADE_DARK_GREEN)
    worksheet.append([])
    for column, value in enumerate(["Grade", "Minimum Total", "Grade Point"], start=1):
        cell = worksheet.cell(19, column, value)
        cell.fill = _fill(GRADE_LIGHT_GREEN)
        cell.font = Font(name="Aptos", bold=True, color=GRADE_DARK_GREEN)
    grade_rows = [
        ("A", details.grade_scale.a_minimum, 5),
        ("B", details.grade_scale.b_minimum, 4),
        ("C", details.grade_scale.c_minimum, 3),
        ("D", details.grade_scale.d_minimum, 2),
        ("E", details.grade_scale.e_minimum, 1),
        ("F", 0, 0),
    ]
    for row_number, values in enumerate(grade_rows, start=20):
        for column_number, value in enumerate(values, start=1):
            worksheet.cell(row_number, column_number, value)

    thin = Side(style="thin", color=BORDER_COLOR)
    for row in worksheet.iter_rows(min_row=3, max_row=25, min_col=1, max_col=3):
        for cell in row:
            if cell.row not in {13, 18}:
                cell.font = Font(
                    name="Aptos",
                    size=10,
                    bold=cell.row == 19,
                    color=TEXT,
                )
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number in list(range(3, 12)) + list(range(14, 17)):
        worksheet.cell(row_number, 1).font = Font(name="Aptos", bold=True, color=MUTED)
    worksheet.column_dimensions["A"].width = 23
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 16
    worksheet.column_dimensions["D"].width = 4
    worksheet.freeze_panes = "A3"
    return worksheet


def _build_final_sheet(
    workbook: Workbook,
    processed: ProcessedFinalResult,
    details: FinalResultDetails,
) -> Any:
    worksheet = workbook.create_sheet("Final Result", 0)
    worksheet.sheet_view.showGridLines = True

    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = "MARK SHEET"
    worksheet["A1"].fill = _fill(TEMPLATE_BEIGE)
    worksheet["A1"].font = Font(
        name="Aptos Display",
        size=18,
        bold=True,
        color=TEMPLATE_BROWN,
    )
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 34

    detail_links = [
        ("Programme", "=Settings!B3"),
        ("Session", "=Settings!B4"),
        ("Semester", "=Settings!B5"),
        ("Course Code", "=Settings!B6"),
        ("Course Title", "=Settings!B7"),
        ("Credit Unit(s)", "=Settings!B8"),
        ("Course Status", "=Settings!B9"),
        ("Lecturer(s)", "=Settings!B10"),
    ]
    for row_number, (label, formula) in enumerate(detail_links, start=3):
        worksheet.cell(row_number, 1, label)
        worksheet.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=8)
        worksheet.cell(row_number, 2, formula)
        worksheet.cell(row_number, 1).font = Font(name="Aptos", bold=True, color=MUTED)
        worksheet.cell(row_number, 2).font = Font(name="Aptos", color=TEXT)
        worksheet.cell(row_number, 2).alignment = Alignment(wrap_text=True)

    # Keep the numeric credit-unit value visible even in spreadsheet viewers
    # that do not calculate formulas when the workbook is first opened.
    worksheet["B8"] = details.credit_units
    worksheet["B8"].number_format = "0"
    worksheet["B8"].alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    _apply_section_header(
        worksheet,
        "A12:H12",
        "Executive Summary",
        fill_color=GRADE_LIGHT_GREEN,
        font_color=GRADE_DARK_GREEN,
    )
    summary_headers = ["Grades", "A", "B", "C", "D", "E", "F", "Total Records"]
    for column_number, value in enumerate(summary_headers, start=1):
        cell = worksheet.cell(13, column_number, value)
        cell.fill = _fill(GRADE_LIGHT_GREEN)
        cell.font = Font(name="Aptos", bold=True, color=GRADE_DARK_GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = 40
    data_end = data_start + len(processed.dataframe) - 1
    data_range_end = max(data_end, data_start)
    worksheet["A14"] = "Count"
    worksheet["A15"] = "Percentage"
    for offset, grade in enumerate("ABCDEF", start=2):
        column_letter = chr(64 + offset)
        worksheet.cell(14, offset, f'=COUNTIF($F${data_start}:$F${data_range_end},"{grade}")')
        worksheet.cell(15, offset, f'=IF($H$14=0,0,{column_letter}14/$H$14)')
        worksheet.cell(15, offset).number_format = "0.0%"
    worksheet["H14"] = f"=COUNTA($B${data_start}:$B${data_range_end})"
    worksheet["H15"] = "=IF(H14=0,0,SUM(B14:G14)/H14)"
    worksheet["H15"].number_format = "0.0%"

    stats = [
        ("A17", "Number of passes", "B17", f'=COUNTIF($H${data_start}:$H${data_range_end},"Pass")'),
        ("C17", "Number of fails", "D17", f'=COUNTIF($H${data_start}:$H${data_range_end},"Fail")'),
        ("E17", "Incomplete records", "F17", f'=COUNTIF($H${data_start}:$H${data_range_end},"Incomplete*")'),
        ("G17", "Pass rate", "H17", "=IF((B17+D17)=0,0,B17/(B17+D17))"),
    ]
    for label_cell, label, value_cell, formula in stats:
        worksheet[label_cell] = label
        worksheet[label_cell].font = Font(name="Aptos", bold=True, color=MUTED)
        worksheet[value_cell] = formula
    worksheet["H17"].number_format = "0.0%"

    chart = BarChart()
    chart.type = "col"
    chart.varyColors = False
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.numFmt = "0%"
    chart.y_axis.delete = True
    chart.y_axis.majorGridlines = None
    chart.height = 4.8
    chart.width = 15.5
    chart.gapWidth = 250
    chart.legend = None
    chart.graphical_properties = GraphicalProperties(noFill=True)
    chart.graphical_properties.line.noFill = True
    chart.add_data(Reference(worksheet, min_col=2, max_col=7, min_row=15), from_rows=True)
    chart.set_categories(Reference(worksheet, min_col=2, max_col=7, min_row=13))
    chart.series[0].graphicalProperties.solidFill = GRADE_GREEN
    chart.series[0].graphicalProperties.line.solidFill = GRADE_GREEN
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showLegendKey = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showPercent = False
    chart.dLbls.showBubbleSize = False
    chart.dLbls.showLeaderLines = False
    chart.dLbls.numFmt = "0.0%"
    chart.dLbls.dLblPos = "outEnd"
    worksheet.add_chart(chart, "B19")

    _apply_section_header(worksheet, "A33:H33", "Lecturer's Remarks")
    worksheet.merge_cells("A34:H34")
    worksheet["A34"] = "=Settings!B11"
    worksheet["A34"].alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.row_dimensions[34].height = 38

    _apply_section_header(
        worksheet,
        "A36:H36",
        "Main Result",
        fill_color=GRADE_LIGHT_GREEN,
        font_color=GRADE_DARK_GREEN,
    )
    worksheet["A37"] = "Maximum CA score"
    worksheet["B37"] = "=Settings!B14"
    worksheet["D37"] = "Maximum exam score"
    worksheet["E37"] = "=Settings!B15"
    worksheet["G37"] = "Maximum total"
    worksheet["H37"] = "=Settings!B14+Settings!B15"
    for cell_ref in ["A37", "D37", "G37"]:
        worksheet[cell_ref].font = Font(name="Aptos", bold=True, color=MUTED)

    headers = list(processed.dataframe.columns)
    for column_number, value in enumerate(headers, start=1):
        cell = worksheet.cell(39, column_number, value)
        cell.fill = _fill(GRADE_LIGHT_GREEN)
        cell.font = Font(name="Aptos", size=10, bold=True, color=GRADE_DARK_GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[39].height = 54

    thin = Side(style="thin", color=BORDER_COLOR)
    for zero_index, row in processed.dataframe.iterrows():
        row_number = data_start + zero_index
        worksheet.cell(row_number, 1, int(row["S/No."]))
        _set_literal(worksheet.cell(row_number, 2), row.iloc[1])
        worksheet.cell(row_number, 3, _safe_excel_value(row["CA Score"]))
        worksheet.cell(row_number, 4, _safe_excel_value(row["Exam Score"]))
        worksheet.cell(
            row_number,
            5,
            f'=IF(OR(C{row_number}="",D{row_number}=""),"",C{row_number}+D{row_number})',
        )
        worksheet.cell(
            row_number,
            6,
            (
                f'=IF(E{row_number}="","",IF(E{row_number}>=Settings!$B$20,"A",'
                f'IF(E{row_number}>=Settings!$B$21,"B",IF(E{row_number}>=Settings!$B$22,"C",'
                f'IF(E{row_number}>=Settings!$B$23,"D",IF(E{row_number}>=Settings!$B$24,"E","F"))))))'
            ),
        )
        worksheet.cell(
            row_number,
            7,
            (
                f'=IF(E{row_number}="","",IF(E{row_number}>=Settings!$B$20,5,'
                f'IF(E{row_number}>=Settings!$B$21,4,IF(E{row_number}>=Settings!$B$22,3,'
                f'IF(E{row_number}>=Settings!$B$23,2,IF(E{row_number}>=Settings!$B$24,1,0))))))'
            ),
        )
        preview_status = str(row["Student Status"])
        if preview_status.startswith("Incomplete"):
            _set_literal(worksheet.cell(row_number, 8), preview_status)
        else:
            worksheet.cell(
                row_number,
                8,
                f'=IF(E{row_number}>=Settings!$B$16,"Pass","Fail")',
            )
        for column_number in range(1, 9):
            cell = worksheet.cell(row_number, column_number)
            cell.fill = _fill(
                GRADE_LIGHT_GREEN if row_number % 2 == 0 else WHITE
            )
            cell.border = Border(bottom=thin)
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(
                horizontal="left" if column_number == 2 else "center",
                vertical="center",
            )
    if len(processed.dataframe):
        table = Table(displayName="FinalResultTable", ref=f"A39:H{data_end}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        status_range = f"H{data_start}:H{data_end}"
        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'$H{data_start}="Pass"'],
                fill=_fill(PALE_GREEN),
                font=Font(color="166534"),
            ),
        )
        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'$H{data_start}="Fail"'],
                fill=_fill(PALE_RED),
                font=Font(color="991B1B"),
            ),
        )
        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'LEFT($H{data_start},10)="Incomplete"'],
                fill=_fill(PALE_AMBER),
                font=Font(color="92400E"),
            ),
        )

    widths = [9, 24, 15, 15, 16, 15, 14, 32]
    for column_number, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + column_number)].width = width
    # The Excel table already supplies its own filter. Adding a worksheet-level
    # filter over the same cells creates two overlapping filter definitions,
    # which desktop Excel treats as damaged workbook content.
    worksheet.print_area = f"A1:H{data_range_end}"
    worksheet.print_title_rows = "39:39"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5
    return worksheet


def build_final_result_workbook(
    processed: ProcessedFinalResult,
    details: FinalResultDetails,
) -> bytes:
    """Create an editable final-result workbook with live Excel formulas."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    _build_settings_sheet(workbook, details)
    _build_final_sheet(workbook, processed, details)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def inspect_final_result_workbook(content: bytes) -> dict[str, Any]:
    """Return compact structural details used by tests and local verification."""

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        result = workbook["Final Result"]
        settings = workbook["Settings"]
        return {
            "sheet_names": workbook.sheetnames,
            "title": result["A1"].value,
            "headers": [result.cell(39, column).value for column in range(1, 9)],
            "first_total_formula": result["E40"].value,
            "first_grade_formula": result["F40"].value,
            "settings_title": settings["A1"].value,
            "settings_credit_units": settings["B8"].value,
            "visible_credit_units": result["B8"].value,
            "visible_credit_units_alignment": result["B8"].alignment.horizontal,
            "chart_count": len(result._charts),
            "table_count": len(result.tables),
            "worksheet_filter": result.auto_filter.ref,
            "freeze_panes": result.freeze_panes,
            "header_fill": result["A39"].fill.fgColor.rgb,
            "header_font_color": result["A39"].font.color.rgb,
            "table_style": result.tables["FinalResultTable"].tableStyleInfo.name,
            "chart_series_fill": (
                result._charts[0]
                .series[0]
                .graphicalProperties.solidFill.srgbClr
            ),
            "chart_values_reference": result._charts[0].series[0].val.numRef.f,
            "chart_data_labels": result._charts[0].dLbls.showVal,
            "chart_data_label_format": result._charts[0].dLbls.numFmt,
            "chart_title": result._charts[0].title,
            "chart_legend": result._charts[0].legend,
            "chart_y_axis_hidden": result._charts[0].y_axis.delete,
            "chart_gridlines": result._charts[0].y_axis.majorGridlines,
            "chart_gap_width": result._charts[0].gapWidth,
            "chart_shows_series_name": result._charts[0].dLbls.showSerName,
            "chart_shows_category_name": result._charts[0].dLbls.showCatName,
            "first_data_row_fill": result["A40"].fill.fgColor.rgb,
            "second_data_row_fill": result["A41"].fill.fgColor.rgb,
        }
    finally:
        workbook.close()
