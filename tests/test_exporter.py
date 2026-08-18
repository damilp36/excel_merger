from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from excel_merger.excel_io import LoadedSheet
from excel_merger.exporter import build_output_workbook, clean_output_filename
from excel_merger.merge import LookupSpec, merge_lookups
from excel_merger.quality import analyze_quality


def test_builds_styled_workbook_with_audit_and_quality_sheets() -> None:
    base = pd.DataFrame({"ID": [1, 2], "Name": ["Ada", "=literal"]})
    lookup = pd.DataFrame({"Key": [1], "Region": ["West"]})
    result = merge_lookups(
        base,
        "ID",
        [
            LookupSpec(
                "lookup.xlsx",
                "Data",
                lookup,
                "Key",
                ["Region"],
                "Lookup",
            )
        ],
    )
    reports = [
        analyze_quality(LoadedSheet("base.xlsx", "Data", 1, base, [])),
        analyze_quality(LoadedSheet("lookup.xlsx", "Data", 1, lookup, [])),
    ]

    content = build_output_workbook(result, reports)
    workbook = load_workbook(BytesIO(content), data_only=False)

    assert workbook.sheetnames == ["Merged Data", "Lookup Audit", "Data Quality"]
    merged = workbook["Merged Data"]
    assert merged.freeze_panes == "A2"
    assert merged.auto_filter.ref == "A1:D3"
    assert merged["B3"].value == "=literal"
    assert merged["B3"].data_type == "s"
    assert merged["C3"].fill.fgColor.rgb.endswith("FEE2E2")
    assert workbook["Lookup Audit"]["A2"].value == "Base rows"


def test_cleans_output_filename() -> None:
    assert clean_output_filename("../../Q3:result.xls") == "Q3_result.xlsx"
    assert clean_output_filename("   ") == "merged_lookup.xlsx"
