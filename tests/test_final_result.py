from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook

from excel_merger.final_result import (
    FinalResultDetails,
    FinalResultWorkbookError,
    GradeScale,
    apply_bulk_score_adjustment,
    build_final_result_workbook,
    import_final_result_workbook,
    inspect_final_result_workbook,
    prepare_edited_final_result,
    prepare_final_result,
)


def make_details() -> FinalResultDetails:
    return FinalResultDetails(
        programme="Example Programme",
        session="2025/2026",
        semester="First",
        course_code="ABC 101",
        course_title="Example Course",
        credit_units=3,
        course_status="Core",
        lecturers="Example Lecturer",
        remarks="Review incomplete records.",
        identifier_heading="Student ID",
        maximum_ca_score=30,
        maximum_exam_score=70,
        grade_scale=GradeScale(),
    )


def test_prepares_grades_statuses_and_incomplete_records() -> None:
    source = pd.DataFrame(
        {
            "Identifier": ["S-1", "S-2", "S-3", "S-4", "S-5"],
            "Assessment": [25, 18, 14, "not a score", 31],
            "Examination": [52, 34, 20, 30, 40],
        }
    )

    processed = prepare_final_result(
        source,
        "Identifier",
        "Assessment",
        "Examination",
        make_details(),
    )

    assert processed.dataframe["Grand Total"].tolist()[:3] == [77.0, 52.0, 34.0]
    assert processed.dataframe["Grade Letter"].tolist()[:3] == ["A", "C", "F"]
    assert processed.dataframe["Student Status"].tolist() == [
        "Pass",
        "Pass",
        "Fail",
        "Incomplete: Invalid CA",
        "Incomplete: Invalid CA",
    ]
    assert processed.grade_counts == {"A": 1, "B": 0, "C": 1, "D": 0, "E": 0, "F": 1}
    assert processed.invalid_ca_count == 2
    assert processed.invalid_exam_count == 0

    edited = processed.dataframe.copy()
    edited.loc[3, "CA Score"] = 20
    recalculated = prepare_edited_final_result(edited, make_details())
    assert recalculated.dataframe["Student Status"].tolist() == [
        "Pass",
        "Pass",
        "Fail",
        "Pass",
        "Incomplete: Invalid CA",
    ]
    assert recalculated.dataframe.loc[3, "CA Score"] == 20
    assert pd.isna(recalculated.dataframe.loc[4, "CA Score"])
    assert recalculated.invalid_ca_count == 1


def test_converts_missing_scores_to_zero_and_names_them_in_status() -> None:
    source = pd.DataFrame(
        {
            "Identifier": ["S-1", "S-2", "S-3"],
            "Assessment": [None, 20, ""],
            "Examination": [50, None, None],
        }
    )
    details = make_details()
    processed = prepare_final_result(
        source,
        "Identifier",
        "Assessment",
        "Examination",
        details,
    )

    assert processed.dataframe["CA Score"].tolist() == [0.0, 20.0, 0.0]
    assert processed.dataframe["Exam Score"].tolist() == [50.0, 0.0, 0.0]
    assert processed.dataframe["Grand Total"].tolist() == [50.0, 20.0, 0.0]
    assert processed.dataframe["Grade Letter"].tolist() == ["C", "F", "F"]
    assert processed.dataframe["Student Status"].tolist() == [
        "Incomplete: Missing CA",
        "Incomplete: Missing Exam",
        "Incomplete: Missing CA and Exam",
    ]
    assert processed.missing_ca_count == 2
    assert processed.missing_exam_count == 2
    assert processed.incomplete_count == 3

    edited = processed.dataframe.copy()
    edited.loc[0, "CA Score"] = 25
    recalculated = prepare_edited_final_result(edited, details)
    assert recalculated.dataframe["Grand Total"].tolist() == [75.0, 20.0, 0.0]
    assert recalculated.dataframe["Student Status"].tolist() == [
        "Pass",
        "Incomplete: Missing Exam",
        "Incomplete: Missing CA and Exam",
    ]
    assert recalculated.missing_ca_count == 1
    assert recalculated.missing_exam_count == 2

    content = build_final_result_workbook(recalculated, details)
    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        result = workbook["Final Result"]
        assert result["C40"].value == 25
        assert result["D41"].value == 0
        assert result["H40"].value.startswith("=IF(E40")
        assert result["H41"].value == "Incomplete: Missing Exam"
        assert result["H42"].value == "Incomplete: Missing CA and Exam"
        assert "Incomplete*" in result["F17"].value
    finally:
        workbook.close()


def test_applies_bulk_marks_by_grade_and_caps_scores() -> None:
    source = pd.DataFrame(
        {
            "Identifier": ["S-1", "S-2", "S-3", "S-4", "S-5"],
            "Assessment": [25, 18, 14, None, 30],
            "Examination": [52, 34, 20, 50, 69],
        }
    )
    details = make_details()
    processed = prepare_final_result(
        source,
        "Identifier",
        "Assessment",
        "Examination",
        details,
    )

    grade_c_adjusted, affected = apply_bulk_score_adjustment(
        processed,
        details,
        "CA Score",
        3,
        grade_letter="C",
    )
    assert affected == 1
    assert grade_c_adjusted.dataframe.loc[1, "CA Score"] == 21
    assert grade_c_adjusted.dataframe.loc[3, "CA Score"] == 0
    assert (
        grade_c_adjusted.dataframe.loc[3, "Student Status"]
        == "Incomplete: Missing CA"
    )

    including_missing, affected = apply_bulk_score_adjustment(
        processed,
        details,
        "CA Score",
        3,
        grade_letter="C",
        include_missing=True,
    )
    assert affected == 2
    assert including_missing.dataframe.loc[3, "CA Score"] == 3
    assert including_missing.dataframe.loc[3, "Student Status"] == "Pass"

    all_exam_adjusted, affected = apply_bulk_score_adjustment(
        processed,
        details,
        "Exam Score",
        5,
    )
    assert affected == 5
    assert all_exam_adjusted.dataframe.loc[4, "Exam Score"] == 70


def test_imports_a_previously_generated_final_result() -> None:
    source = pd.DataFrame(
        {
            "Identifier": ["S-1", "S-2", "S-3"],
            "Assessment": [25, None, "not a score"],
            "Examination": [52, 34, 20],
        }
    )
    details = make_details()
    original = prepare_final_result(
        source,
        "Identifier",
        "Assessment",
        "Examination",
        details,
    )
    content = build_final_result_workbook(original, details)

    imported = import_final_result_workbook(content)

    assert imported.details == details
    assert imported.processed.dataframe["Student ID"].tolist() == [
        "S-1",
        "S-2",
        "S-3",
    ]
    assert imported.processed.dataframe["Student Status"].tolist() == [
        "Pass",
        "Incomplete: Missing CA",
        "Incomplete: Invalid CA",
    ]
    assert imported.processed.missing_ca_count == 1
    assert imported.processed.invalid_ca_count == 1


def test_rejects_a_workbook_that_is_not_a_generated_final_result() -> None:
    workbook = Workbook()
    workbook.active["A1"] = "Not a final result"
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    try:
        import_final_result_workbook(output.getvalue())
    except FinalResultWorkbookError as exc:
        assert "generated final-result workbook" in str(exc)
    else:
        raise AssertionError("Expected an invalid final-result workbook to be rejected.")


def test_rejects_invalid_grade_boundaries() -> None:
    details = make_details()
    invalid_scale = GradeScale(
        a_minimum=60,
        b_minimum=70,
        c_minimum=50,
        d_minimum=45,
        e_minimum=40,
    )
    invalid_details = FinalResultDetails(
        **{**details.__dict__, "grade_scale": invalid_scale}
    )

    assert "descend in order" in " ".join(invalid_details.validation_errors())


def test_builds_formula_driven_final_result_workbook() -> None:
    source = pd.DataFrame(
        {
            "Identifier": ["S-1", "S-2"],
            "Assessment": [25, 18],
            "Examination": [52, 34],
        }
    )
    details = make_details()
    processed = prepare_final_result(
        source,
        "Identifier",
        "Assessment",
        "Examination",
        details,
    )
    content = build_final_result_workbook(processed, details)
    inspection = inspect_final_result_workbook(content)

    assert inspection["sheet_names"] == ["Final Result", "Settings"]
    assert inspection["title"] == "MARK SHEET"
    assert inspection["headers"] == list(processed.dataframe.columns)
    assert inspection["settings_credit_units"] == 3
    assert inspection["visible_credit_units"] == 3
    assert inspection["visible_credit_units_alignment"] == "left"
    assert inspection["first_total_formula"].startswith("=IF(OR(C40")
    assert "Settings!$B$20" in inspection["first_grade_formula"]
    assert inspection["chart_count"] == 1
    assert inspection["table_count"] == 1
    assert inspection["worksheet_filter"] is None
    assert inspection["freeze_panes"] is None
    assert inspection["header_fill"].endswith("E2F0D9")
    assert inspection["header_font_color"].endswith("375623")
    assert inspection["table_style"] == "TableStyleLight1"
    assert inspection["chart_series_fill"] == "548235"
    assert inspection["chart_values_reference"].endswith("$B$15:$G$15")
    assert inspection["chart_data_labels"] is True
    assert inspection["chart_data_label_format"] == "0.0%"
    assert inspection["chart_title"] is None
    assert inspection["chart_legend"] is None
    assert inspection["chart_y_axis_hidden"] is True
    assert inspection["chart_gridlines"] is None
    assert inspection["chart_gap_width"] == 250
    assert inspection["chart_shows_series_name"] is False
    assert inspection["chart_shows_category_name"] is False
    assert inspection["first_data_row_fill"].endswith("E2F0D9")
    assert inspection["second_data_row_fill"].endswith("FFFFFF")

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        assert workbook["Settings"]["B6"].value == "ABC 101"
        assert workbook["Settings"]["B14"].value == 30
        assert workbook["Final Result"]["B3"].value == "=Settings!B3"
        assert workbook["Final Result"]["B8"].value == 3
    finally:
        workbook.close()
